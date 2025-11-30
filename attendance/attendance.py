import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import json
import os

class AttendanceManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Attendance Manager - NeuroKeys: BCI Typing Project")
        self.root.geometry("1920x1080")
        self.root.configure(bg="#f0f0f0")
        
        self.members = []
        self.attendance = {}
        self.current_date = datetime.now().strftime("%Y-%m-%d")
        self.data_file = "Neurokeys - BCI Typing Project.json"
        self.excel_file = "Neurokeys - BCI Typing Project.xlsx"
        self.preview_mode = "txt"
        
        self.load_data()
        self.create_widgets()
        
    def create_widgets(self):
        # Header Frame
        header_frame = tk.Frame(self.root, bg="#218089", height=70)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        header_label = tk.Label(
            header_frame,
            text="📋 Attendance Manager - Project Work A/P",
            font=("Arial", 22, "bold"),
            bg="#218089",
            fg="white"
        )
        header_label.pack(pady=15)
        
        # Controls Frame
        controls_frame = tk.Frame(self.root, bg="white")
        controls_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Date Selection
        date_label = tk.Label(controls_frame, text="📅 Date:", font=("Arial", 10, "bold"), bg="white")
        date_label.grid(row=0, column=0, padx=5, pady=8, sticky="w")
        
        self.date_entry = tk.Entry(controls_frame, width=15, font=("Arial", 11), relief=tk.SOLID, bd=1)
        self.date_entry.insert(0, self.current_date)
        self.date_entry.grid(row=0, column=1, padx=5, pady=8)
        self.date_entry.bind("<KeyRelease>", lambda e: self.update_date())
        
        # Add Members
        member_label = tk.Label(
            controls_frame,
            text="👥 Add Members (comma-separated):",
            font=("Arial", 10, "bold"),
            bg="white"
        )
        member_label.grid(row=0, column=2, padx=15, pady=8, sticky="w")
        
        self.member_input = tk.Entry(controls_frame, width=35, font=("Arial", 10), relief=tk.SOLID, bd=1)
        self.member_input.grid(row=0, column=3, padx=5, pady=8)
        self.member_input.insert(0, "e.g., lucifer, morningstar, radhey")
        
        add_btn = tk.Button(
            controls_frame,
            text="➕ Add Members",
            command=self.add_members,
            bg="#218089",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=12,
            pady=6,
            relief=tk.RAISED,
            bd=2
        )
        add_btn.grid(row=0, column=4, padx=5, pady=8)
        
        # Excel Filename Frame
        filename_frame = tk.Frame(self.root, bg="white")
        filename_frame.pack(fill=tk.X, padx=10, pady=5)
        
        filename_label = tk.Label(filename_frame, text="📁 Excel Filename:", font=("Arial", 10, "bold"), bg="white")
        filename_label.pack(side=tk.LEFT, padx=5)
        
        self.filename_entry = tk.Entry(filename_frame, width=40, font=("Arial", 10), relief=tk.SOLID, bd=1)
        self.filename_entry.pack(side=tk.LEFT, padx=5)
        self.filename_entry.insert(0, "Attendance_Records")
        
        filename_hint = tk.Label(filename_frame, text="(without .xlsx extension)", font=("Arial", 9), bg="white", fg="#666")
        filename_hint.pack(side=tk.LEFT, padx=5)
        
        # Export Buttons
        export_frame = tk.Frame(self.root, bg="white", height=50)
        export_frame.pack(fill=tk.X, padx=10, pady=8)
        export_frame.pack_propagate(False)
        
        excel_btn = tk.Button(
            export_frame,
            text="📊 Update Excel File",
            command=self.update_excel,
            bg="#22c55e",
            fg="white",
            font=("Arial", 11, "bold"),
            padx=18,
            pady=10,
            relief=tk.RAISED,
            bd=2,
            cursor="hand2"
        )
        excel_btn.pack(side=tk.LEFT, padx=8, pady=5)
        
        txt_btn = tk.Button(
            export_frame,
            text="📄 Download TXT",
            command=self.export_txt,
            bg="#3b82f6",
            fg="white",
            font=("Arial", 11, "bold"),
            padx=18,
            pady=10,
            relief=tk.RAISED,
            bd=2,
            cursor="hand2"
        )
        txt_btn.pack(side=tk.LEFT, padx=8, pady=5)
        
        open_excel_btn = tk.Button(
            export_frame,
            text="📂 Open Excel File",
            command=self.open_excel_file,
            bg="#f59e0b",
            fg="white",
            font=("Arial", 11, "bold"),
            padx=18,
            pady=10,
            relief=tk.RAISED,
            bd=2,
            cursor="hand2"
        )
        open_excel_btn.pack(side=tk.LEFT, padx=8, pady=5)
        
        # Stats Frame
        stats_frame = tk.Frame(self.root, bg="white", height=60)
        stats_frame.pack(fill=tk.X, padx=10, pady=8)
        stats_frame.pack_propagate(False)
        
        self.total_label = tk.Label(
            stats_frame,
            text="📊 Total: 0",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="#218089"
        )
        self.total_label.pack(side=tk.LEFT, padx=20, pady=12)
        
        self.present_label = tk.Label(
            stats_frame,
            text="✅ Present: 0",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="#22c55e"
        )
        self.present_label.pack(side=tk.LEFT, padx=20, pady=12)
        
        self.absent_label = tk.Label(
            stats_frame,
            text="❌ Absent: 0",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="#ef4444"
        )
        self.absent_label.pack(side=tk.LEFT, padx=20, pady=12)
        
        self.percentage_label = tk.Label(
            stats_frame,
            text="📈 Attendance: 0%",
            font=("Arial", 12, "bold"),
            bg="white",
            fg="#f59e0b"
        )
        self.percentage_label.pack(side=tk.LEFT, padx=20, pady=12)
        
        # Attendance Section Title
        attendance_label = tk.Label(
            self.root,
            text="✔️ Mark Attendance for " + self.current_date,
            font=("Arial", 13, "bold"),
            bg="#f0f0f0",
            fg="#218089"
        )
        self.attendance_title = attendance_label
        attendance_label.pack(pady=8)
        
        # Canvas with Scrollbar
        canvas_frame = tk.Frame(self.root, bg="white", relief=tk.SUNKEN, bd=1)
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.canvas = tk.Canvas(canvas_frame, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        scrollable_frame = tk.Frame(self.canvas, bg="white")
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.attendance_frame = scrollable_frame
        
        # Mouse wheel scrolling
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel)
        
        # Preview Section with Tabs
        preview_frame = tk.Frame(self.root, bg="#f0f0f0")
        preview_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 10))
        
        preview_label = tk.Label(
            preview_frame,
            text="📊 Real-time Preview",
            font=("Arial", 12, "bold"),
            bg="#f0f0f0",
            fg="#218089"
        )
        preview_label.pack(pady=(5, 5), anchor="w")
        
        # Tab buttons
        tab_frame = tk.Frame(preview_frame, bg="white")
        tab_frame.pack(fill=tk.X, padx=0, pady=(0, 5))
        
        self.txt_tab_btn = tk.Button(
            tab_frame,
            text="📝 TXT Report",
            command=lambda: self.show_preview("txt"),
            bg="#218089",
            fg="white",
            font=("Arial", 10, "bold"),
            relief=tk.RAISED,
            bd=2,
            padx=15,
            pady=5,
            cursor="hand2"
        )
        self.txt_tab_btn.pack(side=tk.LEFT, padx=5, pady=5)
        
        self.excel_tab_btn = tk.Button(
            tab_frame,
            text="📊 Excel Preview",
            command=lambda: self.show_preview("excel"),
            bg="#f0f0f0",
            fg="#333",
            font=("Arial", 10, "bold"),
            relief=tk.RAISED,
            bd=1,
            padx=15,
            pady=5,
            cursor="hand2"
        )
        self.excel_tab_btn.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Preview text area
        self.preview_text = tk.Text(
            preview_frame,
            height=8,
            font=("Courier", 9),
            bg="#f9f9f9",
            relief=tk.SOLID,
            bd=1
        )
        self.preview_text.pack(fill=tk.BOTH, expand=True)
        self.preview_text.config(state=tk.DISABLED)
        
        self.render_attendance()
        self.update_preview()
    
    def show_preview(self, mode):
        """Switch between TXT and Excel preview"""
        self.preview_mode = mode
        
        if mode == "txt":
            self.txt_tab_btn.config(bg="#218089", fg="white", relief=tk.RAISED, bd=2)
            self.excel_tab_btn.config(bg="#f0f0f0", fg="#333", relief=tk.RAISED, bd=1)
        else:
            self.txt_tab_btn.config(bg="#f0f0f0", fg="#333", relief=tk.RAISED, bd=1)
            self.excel_tab_btn.config(bg="#218089", fg="white", relief=tk.RAISED, bd=2)
        
        self.update_preview()
    
    def _on_mousewheel(self, event):
        if event.num == 5 or event.delta < 0:
            self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")
    
    def add_members(self):
        input_text = self.member_input.get().strip()
        if not input_text or input_text == "e.g., Member 1, Member 2, Member 3":
            messagebox.showwarning("⚠️ Warning", "Please enter member names")
            return
        
        new_members = [m.strip() for m in input_text.split(",") if m.strip()]
        added = [m for m in new_members if m not in self.members]
        
        if not added:
            messagebox.showinfo("ℹ️ Info", "All members already exist")
            return
        
        self.members.extend(added)
        self.members.sort()
        
        for member in added:
            if member not in self.attendance:
                self.attendance[member] = {}
        
        self.save_data()
        self.render_attendance()
        self.member_input.delete(0, tk.END)
        self.member_input.insert(0, "e.g., Member 1, Member 2, Member 3")
        self.update_stats()
        self.update_preview()
        messagebox.showinfo("✅ Success", f"Added {len(added)} member(s)")
    
    def render_attendance(self):
        for widget in self.attendance_frame.winfo_children():
            widget.destroy()
        
        if not self.members:
            empty_label = tk.Label(
                self.attendance_frame,
                text="👥 Add members to get started",
                font=("Arial", 12),
                bg="white",
                fg="#999"
            )
            empty_label.pack(pady=30)
            return
        
        for member in self.members:
            is_present = self.attendance.get(member, {}).get(self.current_date) == "P"
            
            member_frame = tk.Frame(self.attendance_frame, bg="#fafafa", relief=tk.FLAT, bd=1, highlightthickness=1, highlightbackground="#e0e0e0")
            member_frame.pack(fill=tk.X, pady=4, padx=5)
            
            var = tk.BooleanVar(value=is_present)
            
            checkbox = tk.Checkbutton(
                member_frame,
                variable=var,
                font=("Arial", 11),
                bg="#fafafa",
                activebackground="#f0f0f0",
                cursor="hand2",
                command=lambda m=member, v=var: self.toggle_attendance(m, v)
            )
            checkbox.pack(side=tk.LEFT, padx=12, pady=10)
            
            label = tk.Label(
                member_frame,
                text=member,
                font=("Arial", 11, "bold"),
                bg="#fafafa",
                fg="#333"
            )
            label.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            
            status_text = "P" if is_present else "A"
            status_color = "#22c55e" if is_present else "#ef4444"
            status_bg = "#e8f5e9" if is_present else "#ffebee"
            
            status_label = tk.Label(
                member_frame,
                text=status_text,
                font=("Arial", 11, "bold"),
                bg=status_bg,
                fg=status_color,
                width=4,
                relief=tk.RAISED,
                bd=1
            )
            status_label.pack(side=tk.RIGHT, padx=12, pady=8)
    
    def toggle_attendance(self, member, var):
        if member not in self.attendance:
            self.attendance[member] = {}
        
        self.attendance[member][self.current_date] = "P" if var.get() else "A"
        self.save_data()
        self.render_attendance()
        self.update_preview()
        self.update_stats()
    
    def update_date(self):
        try:
            self.current_date = self.date_entry.get()
            self.attendance_title.config(text=f"✔️ Mark Attendance for {self.current_date}")
            self.render_attendance()
            self.update_preview()
            self.update_stats()
        except:
            pass
    
    def update_stats(self):
        total = len(self.members)
        present = sum(
            1 for m in self.members
            if self.attendance.get(m, {}).get(self.current_date) == "P"
        )
        absent = total - present
        percentage = int((present / total * 100)) if total > 0 else 0
        
        self.total_label.config(text=f"📊 Total: {total}")
        self.present_label.config(text=f"✅ Present: {present}")
        self.absent_label.config(text=f"❌ Absent: {absent}")
        self.percentage_label.config(text=f"📈 Attendance: {percentage}%")
    
    def update_preview(self):
        self.preview_text.config(state=tk.NORMAL)
        self.preview_text.delete(1.0, tk.END)
        
        if self.preview_mode == "txt":
            self.show_txt_preview()
        else:
            self.show_excel_preview()
        
        self.preview_text.config(state=tk.DISABLED)
    
    def show_txt_preview(self):
        """Show TXT format preview"""
        present_members = [
            m for m in self.members
            if self.attendance.get(m, {}).get(self.current_date) == "P"
        ]
        absent_members = [
            m for m in self.members
            if self.attendance.get(m, {}).get(self.current_date) != "P"
        ]
        
        preview = f"Date: {self.current_date}\n"
        preview += f"Total members: {len(self.members)}\n"
        preview += f"Present ({len(present_members)}): {', '.join(present_members) if present_members else 'None'}\n"
        preview += f"Absent ({len(absent_members)}): {', '.join(absent_members) if absent_members else 'None'}\n\n"
        preview += "Full column preview:\n\n"
        preview += f"{'Member':<30} Status\n"
        preview += "─" * 40 + "\n"
        
        for member in self.members:
            status = self.attendance.get(member, {}).get(self.current_date, "A")
            preview += f"{member:<30} {status}\n"
        
        self.preview_text.insert(1.0, preview)
    
    def show_excel_preview(self):
        """Show Excel format preview - real-time"""
        filename = self.filename_entry.get().strip() or "Neurokeys - BCI Typing Project"
        excel_file = f"{filename}.xlsx"
        
        if not os.path.exists(excel_file):
            preview = f"Excel file will be created at: {excel_file}\n\n"
            preview += "Format Preview:\n"
            preview += "S.n. | Student name | NOVEMBER | DECEMBER | ...\n"
            preview += "─" * 60 + "\n"
            preview += " 1   | Akshansh     |    P     |    P     |\n"
            preview += " 2   | Anuj Dangi   |    A     |    P     |\n"
            preview += "...\n"
            self.preview_text.insert(1.0, preview)
            return
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            preview = f"📊 Real-time Excel Preview: {excel_file}\n"
            preview += "=" * 80 + "\n\n"
            
            # Get headers
            header_row = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                header_row.append(str(val)[:15] if val else "")
            
            # Format header
            preview += "| " + " | ".join(f"{h:15}" for h in header_row) + " |\n"
            preview += "─" * (len(header_row) * 18 + 3) + "\n"
            
            # Get data rows
            for row in range(3, min(ws.max_row + 1, 18)):  # Show up to 15 members
                row_data = []
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row=row, column=col).value
                    row_data.append(str(val)[:15] if val else "")
                preview += "| " + " | ".join(f"{d:15}" for d in row_data) + " |\n"
            
            if ws.max_row > 18:
                preview += f"\n... and {ws.max_row - 17} more rows\n"
            
            self.preview_text.insert(1.0, preview)
            wb.close()
        except Exception as e:
            preview = f"Error reading Excel file:\n{str(e)}\n\nClick 'Update Excel File' to create it."
            self.preview_text.insert(1.0, preview)
    
    def update_excel(self):
        """
            Update Excel file so each MONTH has a single column.
            Row1: Month name (e.g. NOVEMBER)
            Row2: Comma-separated days present in that month (e.g. 27, 28, 29)
            Rows 3+ : S.n., Student name, <MONTH_COLUMN> (status P/A)
        """
        if not self.members:
            messagebox.showwarning("⚠️ Warning", "Add members first")
            return

        try:
            # Get custom filename
            filename = self.filename_entry.get().strip() or "NeuroKeys - BCI Typing Project"
            self.excel_file = f"{filename}.xlsx"

            # Parse current date
            date_obj = datetime.strptime(self.current_date, "%Y-%m-%d")
            month_name = date_obj.strftime("%B").upper()   # e.g. "NOVEMBER"
            day = str(date_obj.day).zfill(2)               # e.g. "27"

            # Load or create workbook
            if os.path.exists(self.excel_file):
                wb = openpyxl.load_workbook(self.excel_file)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Attendance"
                ws['A1'] = "S.n."
                ws['B1'] = "Student name"
                # style header
                for cell in ['A1', 'B1']:
                    ws[cell].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    ws[cell].font = Font(bold=True, color="FFFFFF")
                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

            # Find month column (single column per month)
            month_col = None
            for col in range(3, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and isinstance(header, str) and month_name == header.strip().upper():
                    month_col = col
                    break

            # If month column not found, append at end
            if not month_col:
                month_col = ws.max_column + 1
                ws.cell(row=1, column=month_col, value=month_name)
                ws.cell(row=1, column=month_col).fill = PatternFill(start_color="9966CC", end_color="9966CC", fill_type="solid")
                ws.cell(row=1, column=month_col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=1, column=month_col).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=2, column=month_col, value="")  # initialize row 2

            # Update row 2: maintain a comma-separated list of days for that month
            existing_days = ws.cell(row=2, column=month_col).value
            if existing_days:
                # normalize and add if not present
                days = [d.strip() for d in str(existing_days).split(",") if d.strip()]
                if day not in days:
                    days.append(day)
                    # keep sorted numeric order
                    days = sorted(list(set(days)), key=lambda x: int(x))
                ws.cell(row=2, column=month_col, value=", ".join(days))
            else:
                ws.cell(row=2, column=month_col, value=day)

            # Style row 2
            ws.cell(row=2, column=month_col).fill = PatternFill(start_color="E6D9F3", end_color="E6D9F3", fill_type="solid")
            ws.cell(row=2, column=month_col).font = Font(bold=True, color="000000")
            ws.cell(row=2, column=month_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Update or add members in rows starting from row 3
            # Keep serial numbers consistent with order in self.members
            for idx, member in enumerate(self.members, start=1):
                found_row = None
                # search existing name in column B
                for r in range(3, ws.max_row + 1):
                    val = ws.cell(row=r, column=2).value
                    if val and str(val).strip() == member:
                        found_row = r
                        break

                if not found_row:
                    found_row = ws.max_row + 1 if ws.max_row >= 3 else 3
                    ws.cell(row=found_row, column=1, value=idx)
                    ws.cell(row=found_row, column=2, value=member)
                else:
                    # ensure serial number column is up to date (in case rows moved)
                    ws.cell(row=found_row, column=1, value=idx)

                # Write status for this month (single cell per month column)
                status = self.attendance.get(member, {}).get(self.current_date, "A")
                ws.cell(row=found_row, column=month_col, value=status)

                # Align / border for the row cells
                ws.cell(row=found_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=found_row, column=month_col).alignment = Alignment(horizontal="center", vertical="center")

            # Apply borders and column widths
            max_col = ws.max_column
            for r in range(1, ws.max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if r >= 3:
                        cell.border = Border(
                            left=Side(style='thin', color='CCCCCC'),
                            right=Side(style='thin', color='CCCCCC'),
                            top=Side(style='thin', color='CCCCCC'),
                            bottom=Side(style='thin', color='CCCCCC')
                        )

            # Column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 30
            for col in range(3, max_col + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

            ws.row_dimensions[1].height = 26
            ws.row_dimensions[2].height = 30

            # Save
            wb.save(self.excel_file)
            self.update_preview()
            messagebox.showinfo("✅ Success", f"Excel file updated!\n\n{self.excel_file}\n\nMembers: {len(self.members)}\nDate: {self.current_date}")

        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to update Excel:\n{str(e)}")

            """Update Excel file - merge same month dates into one column"""
            if not self.members:
                messagebox.showwarning("⚠️ Warning", "Add members first")
                return
            
            try:
                # Get custom filename
                filename = self.filename_entry.get().strip()
                if not filename:
                    filename = "Attendance_Records"
                
                self.excel_file = f"{filename}.xlsx"
                
                date_obj = datetime.strptime(self.current_date, "%Y-%m-%d")
                month_num = date_obj.month
                month_name = date_obj.strftime("%B").upper()
                year = date_obj.year
                day = str(date_obj.day).zfill(2)
                
                # Check if file exists and load it
                if os.path.exists(self.excel_file):
                    wb = openpyxl.load_workbook(self.excel_file)
                    ws = wb.active
                else:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Attendance"
                    ws['A1'] = "S.n."
                    ws['B1'] = "Student name"
                    
                    # Format initial headers
                    for cell in ['A1', 'B1']:
                        ws[cell].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                        ws[cell].font = Font(bold=True, color="FFFFFF")
                        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
                
                # Month order mapping
                month_order = {
                    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4,
                    "MAY": 5, "JUNE": 6, "JULY": 7, "AUGUST": 8,
                    "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12
                }
                
                # Find month column - reuse existing month column for same month
                month_col = None
                
                for col in range(3, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if header and isinstance(header, str) and month_name in header.upper():
                        month_col = col
                        break
                
                # If month doesn't exist, find correct position and create it
                if not month_col:
                    insert_pos = 3
                    
                    # Find where to insert: sort months in ascending order
                    for col in range(3, ws.max_column + 1):
                        header = ws.cell(row=1, column=col).value
                        if header and isinstance(header, str):
                            existing_month_num = month_order.get(header.upper(), 13)
                            if existing_month_num < month_num:
                                insert_pos = col + 1
                            elif existing_month_num > month_num:
                                insert_pos = col
                                break
                    
                    # Check if position is valid
                    if insert_pos <= ws.max_column and ws.cell(row=1, column=insert_pos).value:
                        ws.insert_cols(insert_pos)
                        month_col = insert_pos
                    else:
                        month_col = insert_pos if insert_pos <= ws.max_column else ws.max_column + 1
                    
                    ws.cell(row=1, column=month_col, value=month_name)
                    ws.cell(row=2, column=month_col, value="")  # Clear row 2 for month columns
                    
                    cell = ws.cell(row=1, column=month_col)
                    cell.fill = PatternFill(start_color="9966CC", end_color="9966CC", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Update row 2 to show the date range for this month
                row2_val = ws.cell(row=2, column=month_col).value
                if not row2_val:
                    ws.cell(row=2, column=month_col, value=day)
                else:
                    # Merge dates: "27, 28, 29" format
                    existing = str(row2_val).strip()
                    if day not in existing:
                        days_list = [d.strip() for d in existing.split(",")]
                        days_list.append(day)
                        days_list = sorted(list(set(days_list)), key=lambda x: int(x))
                        ws.cell(row=2, column=month_col, value=", ".join(days_list))
                
                cell = ws.cell(row=2, column=month_col)
                cell.fill = PatternFill(start_color="E6D9F3", end_color="E6D9F3", fill_type="solid")
                cell.font = Font(bold=True, color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Update or add members
                for idx, member in enumerate(self.members, 1):
                    # Check if member exists
                    found_row = None
                    for row in range(3, ws.max_row + 1):
                        if ws.cell(row=row, column=2).value == member:
                            found_row = row
                            break
                    
                    if not found_row:
                        found_row = ws.max_row + 1
                        ws.cell(row=found_row, column=1, value=idx)
                        ws.cell(row=found_row, column=2, value=member)
                    
                    # Add attendance for the day
                    status = self.attendance.get(member, {}).get(self.current_date, "A")
                    ws.cell(row=found_row, column=month_col, value=status)
                    
                    # Format cells
                    ws.cell(row=found_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=found_row, column=month_col).alignment = Alignment(horizontal="center", vertical="center")
                
                # Format all cells with borders
                max_col = ws.max_column
                for row in range(1, ws.max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        if row >= 3:
                            cell.border = Border(
                                left=Side(style='thin', color='CCCCCC'),
                                right=Side(style='thin', color='CCCCCC'),
                                top=Side(style='thin', color='CCCCCC'),
                                bottom=Side(style='thin', color='CCCCCC')
                            )
                
                # Column widths
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 25
                for col in range(3, max_col + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
                
                ws.row_dimensions[1].height = 25
                ws.row_dimensions[2].height = 30
                
                wb.save(self.excel_file)
                self.update_preview()  # Update preview in real-time
                messagebox.showinfo("✅ Success", f"Excel file updated!\n\n{self.excel_file}\n\nMembers: {len(self.members)}\nDate: {self.current_date}")
                
            except Exception as e:
                messagebox.showerror("❌ Error", f"Failed to update Excel:\n{str(e)}")
    
    def open_excel_file(self):
        """Open the Excel file with default application (Windows / macOS / Linux)."""
        filename = self.filename_entry.get().strip()
        if not filename:
            filename = "Attendance_Records"
        
        file_to_open = f"{filename}.xlsx"
        
        if not os.path.exists(file_to_open):
            messagebox.showwarning("⚠️ Warning", f"Excel file not found:\n{file_to_open}\n\nClick 'Update Excel File' first!")
            return
        
        try:
            if os.name == 'nt':  # Windows
                os.startfile(file_to_open)
            else:
                # macOS uses 'open', most Linux distros use 'xdg-open'
                if sys.platform == "darwin":
                    cmd = f'open "{file_to_open}"'
                else:
                    cmd = f'xdg-open "{file_to_open}"'
                os.system(cmd)
        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to open file:\n{str(e)}")

    
    def export_txt(self, default_save=False):
        """
        Export current preview as TXT.
        - If default_save=True: save automatically next to Excel file with timestamped name.
        - Otherwise prompts the user with a Save As dialog.
        """
        if not self.members:
            messagebox.showwarning("⚠️ Warning", "Add members first")
            return
        
        present_members = [
            m for m in self.members
            if self.attendance.get(m, {}).get(self.current_date) == "P"
        ]
        absent_members = [
            m for m in self.members
            if self.attendance.get(m, {}).get(self.current_date) != "P"
        ]
        
        content_lines = []
        content_lines.append(f"Date: {self.current_date}")
        content_lines.append(f"Total members: {len(self.members)}")
        content_lines.append(f"Present ({len(present_members)}): {', '.join(present_members) if present_members else 'None'}")
        content_lines.append(f"Absent ({len(absent_members)}): {', '.join(absent_members) if absent_members else 'None'}")
        content_lines.append("")
        content_lines.append("Full column preview:")
        content_lines.append("")
        content_lines.append(f"{'Member':<30} Status")
        content_lines.append("─" * 40)
        
        for member in self.members:
            status = self.attendance.get(member, {}).get(self.current_date, "A")
            content_lines.append(f"{member:<30} {status}")
        
        content = "\n".join(content_lines) + "\n"
        
        if default_save:
            # create default folder next to excel file
            filename = self.filename_entry.get().strip() or "Attendance_Records"
            excel_file = f"{filename}.xlsx"
            folder = os.path.dirname(os.path.abspath(excel_file)) or os.getcwd()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"Attendance_{self.current_date}_{timestamp}.txt"
            file_path = os.path.join(folder, default_name)
            try:
                with open(file_path, 'w', encoding='utf-8', newline='\n') as f:
                    f.write(content)
                messagebox.showinfo("✅ Success", f"TXT file auto-saved:\n{file_path}")
            except Exception as e:
                messagebox.showerror("❌ Error", f"Could not auto-save file:\n{e}")
            return
        
        # interactive save as dialog (original behavior)
        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialfile=f"Attendance_{self.current_date}.txt"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', encoding='utf-8', newline='\n') as f:
                f.write(content)
            messagebox.showinfo("✅ Success", f"TXT file saved!\n\n{file_path}")
        except Exception as e:
            messagebox.showerror("❌ Error", f"Could not save TXT file:\n{e}")

    
    def save_data(self):
        data = {"members": self.members, "attendance": self.attendance}
        with open(self.data_file, 'w') as f:
            json.dump(data, f, indent=2)
    
    def load_data(self):
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r') as f:
                    data = json.load(f)
                    self.members = data.get("members", [])
                    self.attendance = data.get("attendance", {})
            except:
                pass


if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceManager(root)
    root.mainloop()
